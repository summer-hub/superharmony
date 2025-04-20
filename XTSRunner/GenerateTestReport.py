import os
import time
import json
from colorama import Fore
from openpyxl import load_workbook

from config import REPORT_DIR, EXCEL_FILE_PATH
from ReadExcel import read_libraries_from_excel
from ExtractTestDetails import extract_test_details


# 假设这是GenerateTestReport.py中的display_test_tree函数
def display_test_tree(output):
    """显示测试树结构"""
    from ExtractTestDetails import extract_test_details, display_test_details
    
    # 从输出中提取测试结果
    extracted_data = extract_test_details(output)
    test_results = extracted_data["test_results"]
    summary = extracted_data["summary"]
    class_times = extracted_data["class_times"]
    
    # 使用ExtractTestDetails.py中的函数显示测试树
    display_test_details(test_results, summary, class_times)

def update_excel_result(summary, original_name):
    """更新Excel文件中的测试结果"""
    try:
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active

        # 判断测试是否全部通过
        # 修改判断逻辑：当总测试数为0时，应该标记为失败
        if summary["total"] == 0:
            # 如果没有测试被执行，标记为失败
            all_passed = False
            print(f"警告：没有测试被执行 (0 pass, 0 fail)，将结果标记为失败")
        else:
            # 正常情况：只有当没有失败和错误时才标记为通过
            all_passed = (summary["failed"] == 0 and summary["error"] == 0)

        result_value = "pass" if all_passed else "fail"

        # 查找测试结果列的索引
        result_column_index = None
        for i, cell in enumerate(ws[1]):  # 假设第一行是标题行
            if cell.value == "测试结果":
                result_column_index = i
                break

        if result_column_index is None:
            print(Fore.YELLOW + "警告：未找到'测试结果'列，尝试使用第三列" + Fore.RESET)
            result_column_index = 2  # 索引从0开始，第三列索引为2

        # 更新Excel中的测试结果列
        for row_index, row in enumerate(ws.iter_rows(min_row=2), start=2):  # 从第二行开始（跳过标题行）
            # 获取当前行的三方库名称
            library_name = row[0].value
            if library_name == original_name:  # 只更新当前测试的三方库结果
                ws.cell(row=row_index, column=result_column_index + 1).value = result_value
                print(f"已更新 {library_name} 的测试结果为: {result_value}")
                break

        wb.save(EXCEL_FILE_PATH)
        print(f"测试结果已更新到Excel文件")
    except Exception as e:
        print(f"更新Excel文件失败: {str(e)}")


def adjust_test_results_to_match_summary(test_results, summary):
    """调整测试结果以匹配摘要信息"""
    # 计算当前测试结果中的通过和失败数
    current_passed = sum(1 for cls in test_results.values() for test in cls if test.get('status') == 'passed')
    current_total = sum(len(tests) for tests in test_results.values())

    # 如果摘要中有数据但解析结果为空，不创建默认测试类，而是保留原始摘要数据
    if current_total == 0 and summary["total"] > 0:
        print(f"警告：未能从输出中解析出具体的测试类和方法，但摘要显示有 {summary['total']} 个测试")
        return

    # 如果摘要中没有数据但解析结果有数据，使用解析结果更新摘要
    if summary["total"] == 0 and current_total > 0:
        summary["total"] = current_total
        summary["passed"] = current_passed
        summary["failed"] = current_total - current_passed

        print(f"根据解析结果更新了摘要数据: {current_passed}/{current_total} 测试通过")
        return

    # 如果两者都有数据但不一致，只显示警告，不调整测试结果
    if current_total != summary["total"] or current_passed != summary["passed"]:
        print(
            f"警告：解析的测试结果 ({current_passed}/{current_total}) 与摘要 ({summary['passed']}/{summary['total']}) 不一致")


def save_test_json(test_results, summary, class_times, component_name):
    """
    将测试结果保存为JSON文件

    参数:
        test_results: 测试结果字典
        summary: 测试摘要信息
        class_times: 测试类耗时字典
        component_name: 组件名称
    """
    # 创建TestJson目录（如果不存在）
    test_json_dir = os.path.join(os.getcwd(), "TestJson")
    if not os.path.exists(test_json_dir):
        os.makedirs(test_json_dir)

    # 创建一个包含所有信息的字典
    data = {
        "summary": summary,
        "class_times": class_times,
        "test_results": test_results,
        # 添加顶层的 total_time_ms 字段
        "total_time_ms": summary.get('total_time_ms', 0)
    }

    # 生成JSON文件名
    output_file = os.path.join(test_json_dir, f"{component_name}_results.json")

    # 保存为JSON文件
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"测试结果已保存到 {output_file}")


def generate_test_report(test_names, output, original_name):
    """生成HTML格式的测试报告"""
    # 从Excel动态获取库名
    try:
        # 确保original_name是字符串
        if isinstance(original_name, (list, tuple)) and len(original_name) > 0:
            # 如果是列表或元组，使用第一个元素
            library_name = str(original_name[0])
        else:
            library_name = str(original_name) if original_name else ""

        # 尝试从Excel获取库名
        libraries, current_library, _ = read_libraries_from_excel(library_name)

        # 确定组件名 - 如果提供了library_name，使用它，否则使用current_library
        component_name = library_name if library_name else current_library

        # 如果组件名在libraries列表中，使用它，否则使用原始名称
        if component_name not in libraries and original_name:
            component_name = original_name
    except Exception as e:
        print(f"获取库名时出错: {str(e)}")
        # 确保component_name是字符串
        component_name = str(original_name) if original_name else "未知库"

    # 解析测试结果
    extracted_data = extract_test_details(output)
    test_results = extracted_data["test_results"]
    summary = extracted_data["summary"]
    class_times = extracted_data["class_times"]

    # 保存测试结果为JSON
    save_test_json(test_results, summary, class_times, component_name)

    # 更新Excel中的测试结果
    update_excel_result(summary, component_name)

    # 获取当前时间作为报告时间
    current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    # 使用summary中的数据计算测试统计信息
    total_tests = summary["total"]
    total_passed = summary["passed"]
    total_failed = summary["failed"]
    total_skipped = 0  # 如果需要，可以从summary中获取
    total_ignored = summary["ignored"]

    # 计算总执行时间
    total_time_ms = sum(int(test.get('time', '0 ms').replace('ms', '').strip())
                        for cls in test_results.values()
                        for test in cls)
    total_time_s = total_time_ms / 1000.0

    # 准备测试项目HTML
    test_items = []
    overall_status = 'passed' if total_failed == 0 and summary["error"] == 0 else 'failed'

    for test_class, tests in test_results.items():
        # 检查测试类中是否有任何失败的测试
        class_passed = all(test.get('status') == 'passed' for test in tests)
        class_status = 'passed' if class_passed else 'failed'

        # 计算测试类的总执行时间
        class_time_ms = sum(int(test.get('time', '0 ms').replace('ms', '').strip()) for test in tests)

        test_html = []
        for test in tests:
            test_status = test.get('status', 'unknown')
            test_time = test.get('time', '0ms')
            test_name = test.get('name', 'unknown')  # 确保获取测试名称
            error_stack = test.get('error_stack', '')

            # 添加错误堆栈（如果有）
            error_html = f'<div class="error-stack">{error_stack}</div>' if error_stack else ''

            # 修改测试方法HTML，确保测试名称正确显示
            test_html.append(f"""
                            <li class="level test {test_status}">
                                <span>
                                    <div class="time">{test_time}</div>
                                    <div class="status">{test_status}</div>
                                    {test_name}
                                </span>
                                {error_html}
                            </li>""")

        # 修改测试类HTML，确保类状态正确传递
        test_items.append(f"""
                        <li class="level suite {class_status}">
                            <span onclick="toggleTest(this)">
                                <div class="time">{class_time_ms} ms</div>
                                {test_class}
                            </span>
                            <ul style="display: block;">
                                {''.join(test_html)}
                            </ul>
                        </li>""")

    # 构建状态信息字符串 - 这里使用了一些定义在模板中的样式类
    status_info = f'<span class="total">{total_tests} total, </span><span class="passed">{total_passed} passed</span>'
    if total_failed > 0:
        status_info += f', <span class="failed">{total_failed} failed</span>'
    if total_skipped > 0:
        status_info += f', <span class="skipped">{total_skipped} skipped</span>'
    if total_ignored > 0:
        status_info += f', <span class="ignored">{total_ignored} ignored</span>'

    # 直接在代码中定义HTML模板，而不是从外部文件读取
    # 注意：使用双大括号 {{ 和 }} 来表示CSS中的单大括号
    html_template = """<!DOCTYPE html PUBLIC "-//W3C//DTD XHTML 1.0 Strict//EN" "http://www.w3.org/TR/xhtml1/DTD/xhtml1-strict.dtd">
<html xmlns:str="http://exslt.org/strings" xmlns="http://www.w3.org/1999/xhtml">
    <head>
        <META http-equiv="Content-Type" content="text/html; charset=UTF-8">
        <title>Test Results &mdash; {original_name}</title>
        <style type="text/css">
          html {{ height: 100% }}
          body {{
            margin: 0 auto;
            padding: 0;
            text-align: left;
            height: 100%;
            font-family: myriad, arial, tahoma, verdana, sans-serif;
            color: #151515;
            font-size: 90%;
            line-height: 1.3em;
            background-color: #fff;
          }}
          * {{ margin: 0; padding: 0 }}
          .clr {{ clear: both; overflow: hidden; }}
          img {{ border: none }}
          a {{ color: #0046b0; text-decoration: none; }}
          a:hover {{ text-decoration: none; }}
          a:focus, a:active {{ outline: none }}
          .noborder {{ border: none }}
          h1 {{
            color: #151515;
            font-size: 180%;
            line-height: 1.1em;
            font-weight: bold;
          }}
          h2 {{
            color: #393D42;
            font-size: 160%;
            font-weight: normal
          }}
          h3 {{
            font-size: 120%;
            font-weight: bold;
            margin-bottom: .5em
          }}
          h4 {{ font-size: 110%; }}
          h5 {{ font-size: 110%; }}
          span.failed {{ color: #ff0000 }}
          span.error {{ color: #ff0000 }}
          span.passed {{ color: #1d9d01 }}
          span.ignored {{ color: #fff600 }}
          span.skipped {{ color: #fff600 }}
          hr {{ background-color: blue }}
          #container {{ min-width: 30em; }}
          #header {{
            padding: 0;
            position: fixed;
            width: 100%;
            z-index: 10;
            background-color: #c7ceda;
          }}
          #header h1 {{ margin: 1em 3em 1em 1.7em; }}
          #header h1 strong {{ white-space: nowrap; }}
          #header .time {{
            margin-top: 2.2em;
            margin-right: 3.4em;
            float: right;
          }}
          #content {{
            margin: 0;
            padding: .5em 3em .5em 0;
            text-align: left;
            background-color: #fff;
            padding-top: 80px; /* 增加顶部内边距，确保内容不被固定的header遮挡 */
          }}
          #content ul {{
            margin: .4em 0 .1em 2em;
            list-style: none;
          }}
          #content ul li.level {{
            cursor: pointer;
          }}
          #content ul li.level span {{
            display: block;
            font-weight: bold;
            position: relative;
          }}
          #content ul li.level.top {{
            margin-bottom: .3em;
          }}
          #content ul li.level.top > span {{
            padding: .5em 0 .5em 1em;
            font-size: 120%;
            color: #151515;
            background-color: #f2f2f2;
            border-left: solid 10px #93e078;
          }}
          #content ul li.level.top.failed > span {{
            border-left: solid 10px #f02525;
          }}
          #content ul li.level.suite > span {{
            margin-bottom: .8em;
            padding: 0 0 0 .8em;
            display: block;
            font-size: 110%;
            line-height: 1em;
            color: #151515;
            border-left: solid 15px #93e078;
          }}
          #content ul li.level.suite.failed > span {{
            border-left: solid 15px #f02525;
          }}
          #content ul li.level.test {{
            margin-bottom: .5em;
          }}
          #content ul li.level.test > span {{
            padding: .3em 0 .3em 1em;
            color: #0046b0;
            font-size: 100%;
            border-left: solid 6px #93e078;
            border-bottom: solid 1px #dbdbdb;
          }}
          #content ul li.level.test.failed > span {{
            border-left: solid 6px #f02525;
          }}
          #content ul li .time {{
            margin-right: .5em;
            width: 5em;
            text-align: right;
            font-size: 13px;
            color: #151515;
            font-style: normal;
            font-weight: normal;
            position: absolute;
            right: 7em;
            top: 50%;
            transform: translateY(-50%);
          }}
          #content ul li .status {{
            width: 6em;
            font-size: 90%;
            font-style: normal;
            font-weight: normal;
            position: absolute;
            right: 0;
            top: 50%;
            transform: translateY(-50%);
          }}
          /* 设置状态颜色 */
          #content ul li .status {{
            color: #1d9d01; /* 默认为通过颜色 */
          }}

          #content ul li.failed .status {{
            color: #ff0000; /* 失败状态为红色 */
          }}

          /* 确保错误堆栈正确显示 */
          .error-stack {{
            margin-left: 20px;
            padding: 10px;
            background-color: #fff3f3;
            border-left: 3px solid #f02525;
            font-family: monospace;
            white-space: pre-wrap;
            margin-bottom: 10px;
            font-size: 12px;
            color: #333;
            display: block; /* 确保显示 */
          }}

          /* 添加控制按钮样式 */
          .controls {{
            margin: 10px 0;
            padding: 10px;
            background-color: #f5f5f5;
            border-radius: 4px;
            text-align: center;
          }}

          .controls button {{
            margin: 0 5px;
            padding: 5px 10px;
            background-color: #4CAF50;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            font-size: 14px;
          }}

          .controls button:hover {{
            background-color: #45a049;
          }}
        </style>
        <script type="text/javascript">
        // 页面加载完成后立即执行的函数
        document.addEventListener('DOMContentLoaded', function() {{
            // 确保按钮事件绑定
            document.getElementById('expandAllBtn').addEventListener('click', expandAll);
            document.getElementById('collapseAllBtn').addEventListener('click', collapseAll);
            document.getElementById('expandFailedBtn').addEventListener('click', expandFailed);

            // 初始展开所有测试结果
            expandAll();

            // 确保内容区域正好显示在标题栏下方
            var header = document.getElementById('header');
            var content = document.getElementById('content');
            if (header && content) {{
                content.style.paddingTop = (header.offsetHeight + 10) + 'px';
            }}
        }});

        // 切换测试结果显示/隐藏
        function toggleTest(element) {{
            var parent = element.parentNode;
            var children = parent.getElementsByTagName('ul');
            for (var i = 0; i < children.length; i++) {{
                var child = children[i];
                if (child.style.display === 'none') {{
                    child.style.display = 'block';
                }} else {{
                    child.style.display = 'none';
                }}
            }}
        }}

        // 展开所有测试结果
        function expandAll() {{
            console.log("展开全部被点击");
            var elements = document.querySelectorAll('#tree ul');
            for (var i = 0; i < elements.length; i++) {{
                elements[i].style.display = 'block';
            }}
        }}

        // 折叠所有测试结果（保留顶层）
        function collapseAll() {{
            console.log("折叠全部被点击");
            var elements = document.querySelectorAll('#tree ul');
            // 保留第一层，从第二层开始折叠
            for (var i = 0; i < elements.length; i++) {{
                if (i > 0) {{
                    elements[i].style.display = 'none';
                }}
            }}
        }}

        // 只展开失败的测试
        function expandFailed() {{
            console.log("只展开失败项被点击");
            // 先折叠所有
            collapseAll();

            // 然后展开失败的
            var failedElements = document.getElementsByClassName('failed');
            for (var i = 0; i < failedElements.length; i++) {{
                var parent = failedElements[i];
                // 确保父元素也是可见的
                while (parent && parent.tagName) {{
                    if (parent.tagName.toLowerCase() === 'li') {{
                        var uls = parent.getElementsByTagName('ul');
                        for (var j = 0; j < uls.length; j++) {{
                            uls[j].style.display = 'block';
                        }}
                    }}
                    parent = parent.parentNode;
                }}
            }}
        }}
        </script>
    </head>
    <body>
        <div id="container">
            <div id="header">
                <div class="time">{total_time_s:.2f} s</div>
                <h1>
                    {original_name} 测试结果: <strong>{status_info}</strong>
                </h1>
            </div>
            <div id="content">
                <div class="controls">
                    <button id="expandAllBtn">展开全部</button>
                    <button id="collapseAllBtn">折叠全部</button>
                    <button id="expandFailedBtn">只展开失败项</button>
                </div>
                <ul id="tree">
                    <li class="level top {overall_status}">
                        <span onclick="toggleTest(this)"><em class="time">
                                <div class="time">{total_time_s:.2f} s</div>
                            </em>测试结果</span>
                        <ul>
                            {test_items}
                        </ul>
                    </li>
                </ul>
            </div>
            <div id="footer">
                <p>报告生成时间: {current_time}</p>
            </div>
        </div>
    </body>
</html>
"""

    # 使用Python的format方法替换占位符
    html_content = html_template.format(
        current_time=current_time,
        test_items=''.join(test_items),
        total_time_s=total_time_s,
        status_info=status_info,
        overall_status=overall_status,
        original_name=original_name
    )

    # 确保报告目录存在
    if not os.path.exists(REPORT_DIR):
        os.makedirs(REPORT_DIR)

    # 使用组件名生成报告文件
    report_path = os.path.join(REPORT_DIR, f"{component_name}.html")
    
    # 添加调试信息
    print(f"正在生成HTML报告: {report_path}")
    print(f"报告目录: {REPORT_DIR}")
    print(f"组件名: {component_name}")
    
    try:
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        print(f"\n测试报告已生成: {report_path}")
    except Exception as e:
        print(f"写入HTML报告文件时出错: {str(e)}")
        import traceback
        traceback.print_exc()
